import csv
import re
from datetime import date, datetime
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen


PAGE_URL = "https://www.taifex.com.tw/cht/7/vixDaily3MNew"
OUTPUT_CSV = "VIX_Recent_2Months_Data.csv"
OUTPUT_XLSX = "VIX_Recent_2Months_Data.xlsx"

DATE_COL = "\u4ea4\u6613\u65e5\u671f"
TIME_COL = "\u6642\u9593"
VIX_COL = "\u81fa\u6307\u9078\u64c7\u6b0a\u6ce2\u52d5\u7387\u6307\u6578"
AVG_COL = "\u6536\u76e4\u524d1\u5206\u9418\u5e73\u5747\u6307\u6578"
CHANGE_COL = "\u8207\u4e0a\u4e00\u7b46\u5dee\u8ddd%"
YEAR_GAP_COL = "\u96e2\u4eca\u5e74\u9ad8\u4f4e\u9ede\u5dee\u8ddd"
FIELDNAMES = [DATE_COL, TIME_COL, VIX_COL, AVG_COL, CHANGE_COL, YEAR_GAP_COL]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}


def fetch_bytes(url):
    request = Request(url, headers=HEADERS)
    with urlopen(request, timeout=30) as response:
        return response.read()


def decode_taifex_text(data):
    for encoding in ("utf-8-sig", "utf-8", "cp950", "big5"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            pass
    return data.decode("utf-8", errors="replace")


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    value = str(value).strip()
    if " " in value:
        value = value.split(" ", 1)[0]

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    return None


def parse_float(value):
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def get_latest_two_month_files():
    html = decode_taifex_text(fetch_bytes(PAGE_URL))
    matches = re.findall(
        r"https://www\.taifex\.com\.tw/file/taifex/Dailydownload/vix/log2data/(\d{6})new\.txt",
        html,
    )

    months = sorted(set(matches), reverse=True)
    if len(months) < 2:
        raise RuntimeError("Cannot find at least two monthly VIX download files.")

    return [
        (
            month,
            f"https://www.taifex.com.tw/file/taifex/Dailydownload/vix/log2data/{month}new.txt",
        )
        for month in months[:2]
    ]


def parse_month_file(month, url):
    text = decode_taifex_text(fetch_bytes(url))
    rows = []

    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("\u4ea4\u6613\u65e5\u671f") or line.startswith("-"):
            continue

        parts = re.split(r"\s+", line)
        if len(parts) < 4 or not re.fullmatch(r"\d{8}", parts[0]):
            continue

        trade_date = datetime.strptime(parts[0], "%Y%m%d").date()
        rows.append(
            {
                DATE_COL: trade_date.strftime("%Y/%m/%d"),
                TIME_COL: parts[1],
                VIX_COL: parts[2],
                AVG_COL: parts[3],
                "_sort_date": trade_date,
                "_month": month,
            }
        )

    return rows


def row_from_values(values):
    if not values or len(values) < 4:
        return None

    trade_date = parse_date(values[0])
    if trade_date is None:
        return None

    vix_value = parse_float(values[2])
    avg_value = parse_float(values[3])
    return {
        DATE_COL: trade_date.strftime("%Y/%m/%d"),
        TIME_COL: str(values[1]).strip(),
        VIX_COL: "" if vix_value is None else f"{vix_value:.2f}",
        AVG_COL: "" if avg_value is None else f"{avg_value:.2f}",
        "_sort_date": trade_date,
        "_source": "old",
    }


def load_existing_csv(csv_path):
    if not csv_path.exists():
        return []

    rows = []
    with csv_path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        next(reader, None)
        for values in reader:
            row = row_from_values(values)
            if row is not None:
                rows.append(row)
    return rows


def load_existing_xlsx(xlsx_path):
    if not xlsx_path.exists():
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    try:
        workbook = load_workbook(xlsx_path, data_only=True)
    except PermissionError:
        return []
    sheet = workbook.worksheets[0]
    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = row_from_values(values)
        if row is not None:
            rows.append(row)

    return rows


def dedupe_rows(rows):
    merged = {}
    for row in rows:
        merged.setdefault(row["_sort_date"], row)
    return sorted(merged.values(), key=lambda row: row["_sort_date"])


def merge_without_duplicates(existing_rows, downloaded_rows):
    merged = {row["_sort_date"]: row for row in existing_rows}
    added_count = 0

    for row in downloaded_rows:
        if row["_sort_date"] in merged:
            continue
        merged[row["_sort_date"]] = row
        added_count += 1

    return sorted(merged.values(), key=lambda row: row["_sort_date"]), added_count


def add_change_percent(rows_desc):
    for index, row in enumerate(rows_desc):
        current = parse_float(row[VIX_COL])
        previous = parse_float(rows_desc[index + 1][VIX_COL]) if index + 1 < len(rows_desc) else None

        if current is None or previous in (None, 0):
            row[CHANGE_COL] = ""
        else:
            row[CHANGE_COL] = (current / previous) - 1


def add_year_high_low_gaps(rows_desc):
    year_values = {}
    for row in rows_desc:
        value = parse_float(row[VIX_COL])
        if value is None:
            continue
        year_values.setdefault(row["_sort_date"].year, []).append(value)

    year_extremes = {
        year: (max(values), min(values))
        for year, values in year_values.items()
        if values
    }

    for row in rows_desc:
        value = parse_float(row[VIX_COL])
        extremes = year_extremes.get(row["_sort_date"].year)
        if value is None or extremes is None:
            row[YEAR_GAP_COL] = ""
            row["_gap_to_max"] = None
            row["_gap_to_min"] = None
            continue

        year_max, year_min = extremes
        gap_to_max = (year_max - value) / year_max if year_max else None
        gap_to_min = (value - year_min) / year_min if year_min else None
        row["_gap_to_max"] = gap_to_max
        row["_gap_to_min"] = gap_to_min
        row[YEAR_GAP_COL] = f"({gap_to_max:.2%}, {gap_to_min:.2%})"


def save_csv(rows_desc, csv_path):
    with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in rows_desc:
            output = {field: row.get(field, "") for field in FIELDNAMES}
            if isinstance(output[CHANGE_COL], float):
                output[CHANGE_COL] = f"{output[CHANGE_COL]:.2%}"
            writer.writerow(output)


def save_excel(rows_desc, xlsx_path):
    try:
        from openpyxl import Workbook
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Missing openpyxl. Please install it with: pip install openpyxl") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u8cc7\u6599"

    sheet.append(FIELDNAMES)
    for row in rows_desc:
        sheet.append(
            [
                row["_sort_date"],
                row[TIME_COL],
                parse_float(row[VIX_COL]),
                parse_float(row[AVG_COL]),
                row.get(CHANGE_COL, ""),
                row.get(YEAR_GAP_COL, ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = {"A": 14, "B": 16, "C": 24, "D": 24, "E": 18, "F": 26}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy/mm/dd"
    for column in ("C", "D"):
        for cell in sheet[column][1:]:
            cell.number_format = "0.00"
    for cell in sheet["E"][1:]:
        cell.number_format = "0.00%"
    for cell in sheet["F"][1:]:
        cell.alignment = Alignment(horizontal="center")

    max_row = len(rows_desc) + 1
    if len(rows_desc) > 0:
        sheet.conditional_formatting.add(
            f"E2:E{max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="C00000")),
        )
        sheet.conditional_formatting.add(
            f"E2:E{max_row}",
            CellIsRule(operator="lessThan", formula=["0"], font=Font(color="008000")),
        )

        table = Table(displayName="VIXData", ref=f"A1:F{max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    red_font = InlineFont(color="C00000")
    green_font = InlineFont(color="008000")
    black_font = InlineFont(color="000000")
    for row_index, row in enumerate(rows_desc, start=2):
        gap_to_max = row.get("_gap_to_max")
        gap_to_min = row.get("_gap_to_min")
        if gap_to_max is None or gap_to_min is None:
            continue
        sheet.cell(row_index, 6).value = CellRichText(
            TextBlock(black_font, "("),
            TextBlock(red_font, f"{gap_to_max:.2%}"),
            TextBlock(black_font, ", "),
            TextBlock(green_font, f"{gap_to_min:.2%}"),
            TextBlock(black_font, ")"),
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{max_row}"
    workbook.save(xlsx_path)


def update_vix_files():
    base_dir = Path(__file__).resolve().parent
    csv_path = base_dir / OUTPUT_CSV
    xlsx_path = base_dir / OUTPUT_XLSX

    print("Reading old files...")
    existing_rows = dedupe_rows(load_existing_csv(csv_path) + load_existing_xlsx(xlsx_path))
    print(f"Old rows: {len(existing_rows)}")

    print("Reading TAIFEX VIX download list...")
    month_files = get_latest_two_month_files()

    downloaded_rows = []
    for month, url in month_files:
        print(f"Downloading {month}...")
        rows = parse_month_file(month, url)
        print(f"{month} rows: {len(rows)}")
        downloaded_rows.extend(rows)

    if not existing_rows and not downloaded_rows:
        raise RuntimeError("No VIX data to export.")

    merged_rows, added_count = merge_without_duplicates(existing_rows, downloaded_rows)
    duplicate_count = len(downloaded_rows) - added_count
    rows_desc = sorted(merged_rows, key=lambda row: row["_sort_date"], reverse=True)
    add_change_percent(rows_desc)
    add_year_high_low_gaps(rows_desc)

    save_csv(rows_desc, csv_path)
    save_excel(rows_desc, xlsx_path)

    print(f"Added rows: {added_count}")
    print(f"Skipped duplicate rows: {duplicate_count}")
    print(f"Total rows: {len(rows_desc)}")
    print(f"Date range: {merged_rows[0][DATE_COL]} ~ {merged_rows[-1][DATE_COL]}")
    print(f"CSV output: {csv_path}")
    print(f"Excel output: {xlsx_path}")


if __name__ == "__main__":
    try:
        update_vix_files()
    except (HTTPError, URLError, TimeoutError) as exc:
        print(f"Download failed: {exc}")
    except Exception as exc:
        print(f"Run failed: {exc}")
